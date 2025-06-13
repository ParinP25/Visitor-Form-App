import tkinter as tk 
from tkinter import ttk, messagebox  
from openpyxl import Workbook, load_workbook
import os 
import re 

##### This class includes everything  #####
class VisitorFormApp: 
    def __init__(self, root): 
        self.root = root 
        self.root.title("Visitor Sign-Up Form") 
        self.root.attributes('-fullscreen', True)   # makes the window fullscreen   
        
        # adding these to make the form look better  
        self.root.configure(bg="#b4e3f0")        # background color

        self.label_font = ("Segoe UI", 11, "bold")      # fonts 
        self.entry_font = ("Segoe UI", 10) 

        self.main_frame = tk.Frame(self.root, bg="#b4e3f0", padx=30, pady=30, bd=2, relief=tk.RIDGE) 
        self.main_frame.place(relx=0.5, rely=0.5, anchor='center')  

        header = tk.Label(self.main_frame,
                  text="Welcome! Please enter you information",
                  font=("Arial", 20, "bold"),
                  fg="#041D6C", bg="#b4e3f0", pady=20)
        header.grid(row=0, column=0, columnspan=2)

        #header = tk.Label(self.main_frame, text="Visitor Sign-Up Form", font=("Segoe UI", 16, "bold"), bg="#f4f4f4") 
        #header.grid(row=0, columnspan=2, pady=(10, 20)) 
        
        self.create_form_fields()
        self.create_buttons() 

        self.excel_file = "visitor_data.xlsx"
        self.save_to_excel() 

    ##### this will create the layout of the fields and the form 
    def create_form_fields(self): 
        row = 1  

        # Grouping all fields in a LabelFrame called "Personal Information"
        form_section = tk.LabelFrame(self.main_frame,
                             text="Personal Information",
                             font=("Segoe UI", 12, "bold"),
                             bg="#b4e3f0", fg="#041D6C",
                             padx=15, pady=10, labelanchor="n")
        form_section.grid(row=row, column=0, columnspan=2, padx=10, pady=10, sticky="ew")
        row += 1  # Shift rows down so fields don’t overlap the header

        def create_label_entry(label, var_name): 
            nonlocal row 
            tk.Label(form_section, text=label + ":", font=self.label_font, bg="#b4e3f0", anchor="e", justify="right").grid(row=row, column=0, sticky="w", padx=10, pady=5) 
            entry = tk.Entry(form_section, font=self.entry_font, width=50) 
            entry.grid(row=row, column=1, padx=10, pady=5) 
            setattr(self, var_name, entry) 
            row += 1 

        ### Name User Entry
        create_label_entry("Full Name", "name_entry") 

        ### Age User Entry (dropdown)
        tk.Label(form_section, text="Age:", font=self.label_font, bg="#b4e3f0").grid(row=row, column=0, sticky="w", padx=10, pady=5) 
        self.age_var = tk.StringVar() 
        self.age_combo = ttk.Combobox(form_section, textvariable=self.age_var, font=self.entry_font, width=47, state="readonly") 
        self.age_combo['values'] = [str(i) for i in range(0, 101)] 
        self.age_combo.grid(row=row, column=1, padx=10, pady=5) 
        row += 1  

        ### Phone Number User Entry 
        create_label_entry("Phone Number", "phone_entry")

        ### Email User Entry 
        create_label_entry("Email", "email_entry") 

        ### Living Address User Entry 
        create_label_entry("Address", "address_entry")  

        ### Place of Origin User Entry 
        create_label_entry("Place of Origin (e.g., Alberta, Canada)", "origin_entry")

        ### Employment Status User Entry (dropdown)
        tk.Label(form_section, text="Employment Status:", font=self.label_font, bg="#b4e3f0").grid(row=row, column=0, sticky="w", padx=10, pady=5) 
        self.employment_var = tk.StringVar() 
        self.employment_combo = ttk.Combobox(form_section, textvariable=self.employment_var, font=self.entry_font, width=47, state="readonly") 
        self.employment_combo['values'] = ["Employed", "Unemployed", "Retired", "Student", "Other"] 
        self.employment_combo.grid(row=row, column=1, padx=10, pady=5)  
        row += 1 

        ### Immigration Status User Entry 
        tk.Label(form_section, text="Immigration Status:", font=self.label_font, bg="#b4e3f0").grid(row=row, column=0, sticky="w", padx=10, pady=5) 
        self.immigration_var = tk.StringVar() 
        self.immigration_combo = ttk.Combobox(form_section, textvariable=self.immigration_var, font=self.entry_font, width=47, state="readonly") 
        self.immigration_combo['values'] = ["Citizen", "Permanent Resident", "Work Permit", "Visitor", "Other"] 
        self.immigration_combo.grid(row=row, column=1, padx=10, pady=5) 
        row += 1 

        ### Disclaimer Options User Entry  
        tk.Label(self.main_frame, text="Disclaimer", font=self.label_font, bg="#b4e3f0").grid(row=row, column=0, sticky="w", padx=10, pady=10)  

        disclaimer_text = ("Disclaimer: Please read and answer below:\n" 
                           "The information collected in this form is used solely for the purpose of tracking building entry and "
            "ensuring the safety and security of all visitors and staff. Your data will be stored securely and will not "
            "be shared with third parties without your consent, unless required by law.\n\n"
            "We may also use the email provided to send you information about events happening throughout the year. "
            "You can opt out of these communications at any time.\n\n"
            "By submitting this form, you acknowledge and agree to the collection and use of your information as described."
        ) 
        tk.Label(self.main_frame, text=disclaimer_text, wraplength=600, justify="left", font=("Segoe UI", 9), bg="#b4e3f0").grid(row=row, column=1, padx=10, pady=10, sticky="w")  
        row += 1 

        self.disclaimer_var = tk.StringVar()   
        self.disclaimer_var.set(None) 

        tk.Radiobutton(self.main_frame, text="I agree and want to receive updates", 
                       variable=self.disclaimer_var, value="Agree + With Updates", bg="#b4e3f0").grid(row=row, column=1, sticky="w", padx=10)  
        row += 1 

        tk.Radiobutton(self.main_frame, text="I agree but do NOT want to receive updates", 
                       variable=self.disclaimer_var, value="Agree Only + No Updates", bg="#b4e3f0").grid(row=row, column=1, sticky="w", padx=10)  
        row += 1 
        
    ##### Buttons Layout  
    def create_buttons(self):  
        button_frame = tk.Frame(self.main_frame, bg="#b4e3f0")   
        button_frame.grid(row=99, columnspan=2, pady=20)  
            
        tk.Button(button_frame, text="Review & Submit", command=self.review_form, font=self.label_font, width=15).pack(side="left", padx=10)  
        tk.Button(button_frame, text="Clear All", command=self.clear_all, font=self.label_font, width=15).pack(side="left", padx=10)  
        tk.Button(button_frame, text="Exit", command=self.root.quit, font=self.label_font, width=15).pack(side="left", padx=10) 


    ##### Clear All, clears the entire form 
    def clear_all(self): 
        self.name_entry.delete(0, tk.END)
        self.age_var.set("")
        self.phone_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
        self.address_entry.delete(0, tk.END)
        self.origin_entry.delete(0, tk.END)
        self.employment_var.set("")
        self.immigration_var.set("")
        self.disclaimer_var.set(None)

    ##### Function for Error Handling, to make sure user inputs are valid 
    def valid_inputs(self): 
        if not self.name_entry.get().strip():  
            messagebox.showerror("Missing", "Please enter your full name.") 
            return False 
        if not self.age_var.get(): 
            messagebox.showerror("Missing", "Please select your age.") 
            return False 
        if not self.phone_entry.get().strip().isdigit():
            messagebox.showerror("Invalid", "Please enter a valid phone number.") 
            return False 
        if "@" not in self.email_entry.get() or "." not in self.email_entry.get(): 
            messagebox.showerror("Invalid", "Please enter a valid email address.") 
            return False 
        if not self.address_entry.get().strip(): 
            messagebox.showerror("Missing", "Please enter your residential address.")
            return False
        origin = self.origin_entry.get().strip()
        if not origin or "," not in origin: 
            messagebox.showerror("Invalid", "Please enter your place of origin as 'Province, Country'.") 
            return False 
        if not self.employment_var.get(): 
            messagebox.showerror("Missing", "Please select your employment status.") 
            return False 
        if not self.immigration_var.get(): 
            messagebox.showerror("Missing", "Please select your immigration status.")
            return False
        if not self.disclaimer_var.get(): 
            messagebox.showerror("Missing", "Please select an option.") 
            return False 
        return True 
    
    ##### Function for form review 
    def review_form(self): 
        if not self.valid_inputs(): 
            return 
        
        ### Create Review Message 
        info = ( 
            f"Full Name: {self.name_entry.get()}\n"
            f"Age: {self.age_var.get()}\n"
            f"Phone Number: {self.phone_entry.get()}\n" 
            f"Email Address: {self.email_entry.get()}\n"
            f"Residential Address: {self.address_entry.get()}\n"
            f"Place of Origin: {self.origin_entry.get()}\n"
            f"Employment Status: {self.employment_var.get()}\n"
            f"Immigration Status: {self.immigration_var.get()}\n"
            f"Disclaimer Choice: {self.disclaimer_var.get()}\n"
        ) 
        if messagebox.askokcancel("Review your information", info + "\n\nSubmit this information?"): 
            self.save_to_excel()
            self.clear_all()
            messagebox.showinfo("Thank You", "Your information has been submitted successfully!")  

    ##### this function will save the excel file and create a row for each user 
    def save_to_excel(self):  
        # Don't save if required fields are empty — safeguard
        if not all([
            self.name_entry.get().strip(),
            self.age_var.get(),
            self.phone_entry.get().strip(),
            self.email_entry.get().strip(),
            self.address_entry.get().strip(),
            self.origin_entry.get().strip(),
            self.employment_var.get(),
            self.immigration_var.get(),
            self.disclaimer_var.get()
        ]):
            return # Exit early — don't write anything

        # If file doesn't exist, create it with headers
        if not os.path.exists(self.excel_file): 
            wb = Workbook()
            ws = wb.active 
            ws.append(["Name", "Age", "Phone Number", "Email Address", "Residential Address", 
                        "Place of Origin", "Employment Status", "Immigration Status", "Disclaimer Choice"]) 
            wb.save(self.excel_file) 

        wb = load_workbook(self.excel_file)
        ws = wb.active 
        ws.append([ 
            self.name_entry.get(), 
            self.age_var.get(), 
            self.phone_entry.get(), 
            self.email_entry.get(), 
            self.address_entry.get(), 
            self.origin_entry.get(), 
            self.employment_var.get(), 
            self.immigration_var.get(), 
            self.disclaimer_var.get()
        ]) 
        wb.save(self.excel_file) 

##### run the program 
if __name__ == "__main__": 
    root = tk.Tk() 
    app = VisitorFormApp(root) 
    root.mainloop() 
